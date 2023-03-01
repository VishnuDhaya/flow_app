import json
import sys
import pandas as pd
from fuzzywuzzy import process
import numpy as np
from openpyxl import load_workbook
from datetime import datetime as dt
import pathlib
import logging

IDENTIFIER_COLUMNS = ['agent_id', 'agent_msisdn']
logger = None

def get_commission_headers(df):
    headers = df.columns
    commission_headers = []
    for header in headers:
        header = str(header)
        try:
            dt.strptime(header, '%b-%Y')
        except ValueError:
            continue
        commission_headers.append(header)
    if not commission_headers:
        raise Exception ('Commission Headers not set')

    return commission_headers

def get_sheet_names(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def rename(df):
    columns = df.columns
    msisdn = process.extractOne('agent_msisdn', columns)[0] 
    agent_id = process.extractOne('AGENT ID', columns)[0] 
    df.rename(columns={
        msisdn: 'agent_msisdn',
        agent_id: 'agent_id',
    }, inplace=True)
    return df

def drop_unwanted_columns(df, columns):
    df = df.loc[:,df.columns.isin(columns)]
    return df

def reorder_columns(df):
    cols = df.columns.to_list()
    move_to_start_columns = IDENTIFIER_COLUMNS
    for move_to_start_column in move_to_start_columns:
        cols.insert(0, cols.pop(cols.index(move_to_start_column)))
    df = df.loc[:, cols]
    return df

def clean_dataframe(df):
    df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    # nan_values = [r'^\s*$', r'-', r'#############', r'############']
    # for nan_value in nan_values:
    #     df = df.replace(nan_value, np.nan, regex=True)
    return df

def drop_n_check(df, commission_headers):
    required_columns = IDENTIFIER_COLUMNS + commission_headers
    df = drop_unwanted_columns(df, required_columns)
    header = df.columns.to_list()
    remaining = list(set(IDENTIFIER_COLUMNS).difference(set(header)))
    if remaining: raise Exception(f"The required columns: {IDENTIFIER_COLUMNS} are not present in statement: {header}")
    return df

def merge_all_sheets(file_name, sheet_names):
    final_df = pd.DataFrame()

    all_commission_headers = []
    for sheet_name in sheet_names:
        logger.warning(f"-----------------------------{sheet_name}---------------------------------")
        df = pd.read_excel(file_name, sheet_name=sheet_name, thousands=',')
        
        df = clean_dataframe(df)
        df = rename(df)

        commission_headers = get_commission_headers(df)
        for commission_header in commission_headers:
            if commission_header in all_commission_headers:
                df.drop([commission_header], axis=1, inplace=True, errors='ignore')
            else:
                df[commission_header] = pd.to_numeric(df[commission_header], errors='coerce')
                all_commission_headers.append(commission_header)

        df = drop_n_check(df, commission_headers)
        final_df = df if final_df.empty else pd.merge(df, final_df, how="outer", left_on=IDENTIFIER_COLUMNS, right_on=IDENTIFIER_COLUMNS)

    final_df = reorder_columns(final_df)
    final_df['avg'] = final_df.iloc[:, 2:].mean(axis=1)
    final_df['avg'] = final_df.iloc[:, 2:].max(axis=1)
    final_df = final_df.groupby('agent_msisdn').first()

    return final_df

def transform_sheet(data):
    
    file_name = data['file_name']
    base_path = data['base_path']
    base_path = pathlib.Path(base_path)
    storage_file_path = f'storage/data/comms/{file_name}'
    file_path = f'{storage_file_path}.xlsx'
    abs_file_path = base_path.joinpath(file_path)

    sheet_names = get_sheet_names(abs_file_path)
    final_df = merge_all_sheets(abs_file_path, sheet_names)
    
    result_file = f'{storage_file_path}.csv'
    result_file_path = base_path.joinpath(result_file)
    
    final_df.to_csv(result_file_path)

def main(data):

    base_path = data['base_path']
    now = dt.now().strftime('%Y-%m-%d')
    logging.basicConfig(filename="{}/storage/logs/{}_comms_import_log.log".format(base_path, now),
                            filemode="a",
                            format='%(asctime)s %(message)s',
                            level=logging.WARNING)
    global logger
    logger = logging.getLogger()
    logger.warning('=====================MAIN FUNCTION STARTED=====================')
    transform_sheet(data)
    


if __name__ == '__main__':
    data = json.loads(sys.argv[1])
    main(data)